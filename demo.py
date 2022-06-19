import streamlit as st
import pandas as pd
import json 
import numpy as np
from openpyxl import load_workbook, Workbook
from urllib.request import urlopen
import requests, json
import xlsxwriter
from zipfile import ZipFile
from io import BytesIO
import base64


with open('style.css') as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

def style_table(v):
    if v < 4:
        return 'color:green;'
    elif v > 45:
        return 'color:red;'
    else:
        return None
@st.cache(allow_output_mutation=True)
def convert_df(df):
     # IMPORTANT: Cache the conversion to prevent computation on every rerun
     return df.to_csv(index=False)
@st.cache(allow_output_mutation=True)
def get_survey_data(survey_db):
    survey_db = 'https://raw.githubusercontent.com/jelisavetaM/VS_module/main/220437.xlsx'
    return pd.read_excel(survey_db)
@st.cache
def get_vs_data(vs_db_files):
    df_vs = pd.DataFrame()
    vs_db_files = ['https://raw.githubusercontent.com/jelisavetaM/VS_module/main/Report%20Products%20-%202022044_vs_cell1.csv','https://raw.githubusercontent.com/jelisavetaM/VS_module/main/Report%20Products%20-%202022044_vs_cell2.csv', 'https://raw.githubusercontent.com/jelisavetaM/VS_module/main/Report%20Products%20-%202022044_vs_cell3.csv']
    for file in vs_db_files:
        df = pd.read_csv(file, delimiter=";" , keep_default_na=False)
        df_vs = df_vs.append(df)
    df_vs = df_vs[df_vs['USER ID'] != '']
    df_vs['CONSIDERATIONS'] = np.where(df_vs['CONSIDERATIONS'] == 'NULL', 0, df_vs['CONSIDERATIONS'])
    df_vs['QUANTITY'] = np.where(df_vs['QUANTITY'] == 'NULL', 0, df_vs['QUANTITY'])
    df_vs = df_vs.astype({'CONSIDERATIONS':'int', 'QUANTITY':'int'})
    df_vs['CONSIDERATIONS_BINARY'] = df_vs['CONSIDERATIONS'].apply (lambda x: 1 if x > 0 else 0)
    df_vs['PENETRATION_BINARY'] = df_vs['QUANTITY'].apply (lambda x: 1 if x > 0 else 0)
    df_vs.replace(to_replace = '', value = 'NOT DEFINED', inplace = True)
    df_vs.replace(to_replace = 'NULL', value = 'NO SHOPPING', inplace = True)
    #trenutno, dok se ne vidi sta je bug sa money spent na VS platformi
    df_vs = df_vs[ (df_vs['PENETRATION_BINARY'] == 1) & (df_vs['MONEY SPENT'] != 'NO SHOPPING') ]
    df_vs = df_vs.astype({'MONEY SPENT':'float', 'PRICE':'float'})
    # st.write(df_vs)
    return df_vs
@st.cache(allow_output_mutation=True)
def get_datamap(datamap_json_file):
    datamap = {}
    questions_label_text = []
    #datamap_json = json.load(datamap_json_file)
    url = requests.get("https://raw.githubusercontent.com/jelisavetaM/VS_module/main/datamap.json")
    text = url.text
    datamap_json = json.loads(text)
    #response = urlopen("https://github.com/jelisavetaM/VS_module/blob/main/datamap.json")
    #datamap_json = json.loads(response.read())
    for var in datamap_json["variables"]:
        q_title = var["label"]
        answers = {}
        if "value" in var:
            answers[0] = "NO TO: " + var["rowTitle"]
            answers[var["value"]] = var["rowTitle"]
        elif "values" in var:
            for val in var["values"]:
                answers[val["value"]] = val["title"]
        q_json = {
            "text" : var["title"],
            "type" : var["type"],
            "vgroup" : var["vgroup"],
            "answers" : answers
        
        }
        datamap[q_title] = q_json
        questions_label_text.append(q_title + "->" + var["title"])
    return [datamap,questions_label_text]   
@st.cache(allow_output_mutation=True)
def get_df_with_answer_labels(df,vars_arr):
    global datamap
    if vars_arr == "ALL":
        df_return = df
    else:
        df_return = df[vars_arr]
    
    for col in df_return.columns:
        labels = list(set(df_return[col].tolist()))
        for lab in labels:
            if datamap[col]["answers"] and lab in datamap[col]["answers"]:
                df_return[col] = df_return[col].replace(lab, datamap[col]["answers"][lab])
    
    return df_return


def format_splits(splits):
    global uuid_and_split
    
    for lvl in splits:
        splits_short = []
        if lvl == "1":
            splits_short.append("CELL")
        for split in splits[lvl]:
            if split.split("->")[0] not in splits_short:
                splits_short.append(split.split("->")[0])
        splits[lvl] = splits_short
    splits_final = {"1" : splits["1"]}

    uuid_and_split = list(np.concatenate([splits["1"].copy(),splits["2"].copy(),splits["3"].copy()]).flat)#ovde treba flatten za sva 3 nivoa splita
    uuid_and_split.append("uuid")
    
    if len(splits["2"]) > 0:
        lvl2 = []
        for s1 in splits["1"]:
            for s2 in splits["2"]:
                if s2!=s1:
                    lvl2.append([s1,s2])
        splits_final["2"] = lvl2
    else:
        splits_final["2"] = []

    if len(splits["3"]) > 0:
        lvl3 = []
        for lvl2 in splits_final["2"]:
            for s3 in splits["3"]:
                pom_niz = lvl2.copy()
                if s3 not in pom_niz:
                    pom_niz.append(s3)
                    lvl3.append(pom_niz)
        splits_final["3"] = lvl3
    else:
        splits_final["3"] = []

    return splits_final

def format_tables(workbook, worksheet, number_of_sheet_rows):
    format_procenti = workbook.add_format({'num_format': '0%'})


    worksheet.conditional_format("$B$1:$QQ$%d" % (number_of_sheet_rows),
                                {"type": "formula",
                                "criteria": '=INDIRECT("D"&ROW())="Penetration on total sample" ',
                                "format": format_procenti
                                })
    worksheet.conditional_format("$B$1:$QQ$%d" % (number_of_sheet_rows),
                                {"type": "formula",
                                "criteria": '=INDIRECT("D"&ROW())="Consideration on total sample"',
                                "format": format_procenti
                                })
    worksheet.conditional_format("$B$1:$QQ$%d" % (number_of_sheet_rows),
                                {"type": "formula",
                                "criteria": '=INDIRECT("D"&ROW())="Share of Total Units"',
                                "format": format_procenti
                                })
    worksheet.conditional_format("$B$1:$QQ$%d" % (number_of_sheet_rows),
                                {"type": "formula",
                                "criteria": '=INDIRECT("D"&ROW())="Share of Total Value"',
                                "format": format_procenti
                                })



def get_measure_df(measure, level, split):
	global shoppingMergedData, data_survey
	definition = {
	'Consideration on total sample': 
		{'filters' : "CONSIDERATIONS_BINARY",
		'data' : "USER ID",
		'aggfunction' : "nunique",
		'base' : "fullBase"}
	,
	'Penetration on total sample': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "USER ID",
		'aggfunction' : "nunique",
		'base' : "fullBase"}
	,
	'Consideration on considerers': 
		{'filters' : "CONSIDERATIONS_BINARY",
		'data' : "USER ID",
		'aggfunction' : "nunique",
		'base' : "shoppers"}
	,
	'Penetration on shoppers': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "USER ID",
		'aggfunction' : "nunique",
		'base' : "fullBase"}
	,
	'Unit Buy Rate (Units per Buyer)': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "QUANTITY",
		'aggfunction' : "mean",
		'base' : 1}
	,
	'Value Buy Rate(Value per Buyer)': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "MONEY SPENT",
		'aggfunction' : "mean",
		'base' : 1}
	,
	'Total Units': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "QUANTITY",
		'aggfunction' : "sum",
		'base' : 1}
	,
	'Total Value': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "MONEY SPENT",
		'aggfunction' : "sum",
		'base' : 1}
	,
	'Share of Total Units': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "QUANTITY",
		'aggfunction' : "sum",
		'base' : 1}
	,
	'Share of Total Value': 
		{'filters' : "PENETRATION_BINARY",
		'data' : "MONEY SPENT",
		'aggfunction' : "sum",
		'base' : 1}
	}
	# st.write(shoppingMergedData)
	# st.stop()
	if measure == "Total Units" or measure == "Total Value" or measure == "Unit Buy Rate (Units per Buyer)" or measure == "Value Buy Rate (Units per Buyer)":
		kpi = shoppingMergedData[shoppingMergedData[definition[measure]['filters']] == 1].pivot_table(definition[measure]['data'], index=level, columns=split, aggfunc=definition[measure]['aggfunction'],  margins=True, margins_name='Total').fillna(0).round(0)
	elif measure == "Share of Total Units" or measure == "Share of Total Value":
		kpiTemp = shoppingMergedData[shoppingMergedData[definition[measure]['filters']] == 1].pivot_table(definition[measure]['data'], index=level, columns=split, aggfunc=definition[measure]['aggfunction'],  margins=True, margins_name='Total')
		kpi = (kpiTemp/kpiTemp.sum()).fillna(0).round(4)
		# kpi = ((kpiTemp/kpiTemp.sum())*100).fillna(0).round(0).astype(int).astype(str) + '%'
	else:
		kpi = shoppingMergedData.pivot_table(definition[measure]['data'], index=level, columns=split, aggfunc=definition[measure]['aggfunction'],  margins=True, margins_name='Total')
		sampleSizes = data_survey[split].value_counts()
		sampleSizes['Total'] = sampleSizes.sum()
		if measure != "Unit Buy Rate (Units per Buyer)" and measure != "Value Buy Rate (UnitsValue per Buyer)":
			kpi = kpi.div(sampleSizes).fillna(0).round(4)
			# kpi = ((kpi.div(sampleSizes))*100).fillna(0).round(0).astype(int).astype(str) + '%'
		N = pd.DataFrame(data = [sampleSizes], index = ['Sample size'], columns=kpi.columns)
		kpi = pd.concat([N, kpi])
	
    
	kpi = kpi.reset_index()
	kpi.rename(columns = {'index':level}, inplace = True)
	return kpi
    
def splitEngine(measures, splitScheme, levels):
    global shoppingMergedData

    tables = []
    tables_by_measure = {}
    arrays_by_measure = [[],[]]
    for level in levels:
        dfAll = pd.DataFrame()
        # arrays = [
                # ["bar", "bar", "baz", "baz", "foo", "foo", "qux", "qux"],
                # ["one", "two", "one", "two", "one", "two", "one", "two"]
        # ]
        arrays = [[],[]]
        
        sublevels = levels[level]

        if level not in tables_by_measure:
            tables_by_measure[level] = {}

        for measure in measures:
            df_splits = pd.DataFrame()
            sp_arr = ["", "", ""]
            for split in splitScheme:
                df = get_measure_df(measure,level,split)

                # st.write(df.astype(str))
                try:
                    df = df[df[level].isin(sublevels)]
                except:
                    st.error("Calculation get_measure_df failed for measure: " + measure)
                    st.write(df.astype(str))
                    st.stop()
                
                # st.stop()
                
                df.insert(0, 'level', level)
                df = df.rename(columns={level: "sublevel","Total" : "Total_" + str(split)})
                
                for x in range(0,(df.shape[1]-2)):
                    sp_arr.append(split)
                
                if df_splits.empty:
                    df.insert(2, 'measurment', measure)
                    df_splits = df
                else:
                    df_splits = pd.merge(df_splits, df, how='left', on=["level","sublevel"])

            # st.write(df_splits.astype(str))

            dfAll = pd.concat([dfAll,df_splits])
            if len(arrays[0])==0:
                arrays[0] = sp_arr
            if len(arrays_by_measure[0])==0:
                arrays_by_measure[0] = sp_arr

            if measure not in tables_by_measure[level]:
                tables_by_measure[level][measure] = pd.DataFrame()
            
            try:
                tables_by_measure[level][measure] = pd.concat([tables_by_measure[level][measure], df_splits])
            except:
                st.write("aaaaaaaaaaaaa")
                st.write(tables_by_measure[measure])
                st.write(df_splits)
                st.stop()
            

        dfAll = dfAll.sort_values(by=['sublevel'])
        
        # OVDE TREBA DODATI DA UBACI SAMPLE SIZE
        # sampleSizes = data_survey[split].value_counts()
        # sampleSizes['Total'] = sampleSizes.sum()
        # N = pd.DataFrame(data = [sampleSizes], index = ['Sample size'], columns=dfAll.columns)
        # dfAll = pd.concat([N, dfAll])
        
        arrays[1] = list(dfAll.columns)
        tuples = list(zip(*arrays))
        multi_column_names = pd.MultiIndex.from_tuples(tuples, names=["Var Name", "Var Label"])
        
        dfAll.columns = multi_column_names
        dfAll.reset_index(drop=True, inplace=True)
        tables.append(dfAll)
        st.info(level)
        # st.write(dfAll.astype(str))


    for level in tables_by_measure:
        for measure in tables_by_measure[level]:
            arrays_by_measure[1] = list(tables_by_measure[level][measure].columns)
            tuples = list(zip(*arrays_by_measure))
            multi_column_names = pd.MultiIndex.from_tuples(tuples, names=["Var Name", "Var Label"])
            tables_by_measure[level][measure].columns = multi_column_names
            tables_by_measure[level][measure].reset_index(drop=True, inplace=True)


    # st.write(tables_by_measure)
    return [tables,tables_by_measure]

#v2
def splitEngine2(measures, splitScheme, levels):
    global shoppingMergedData
    dfAll_tables = {}
    for level_number, splits in splitScheme.items(): 
        table = pd.DataFrame()
        arrays = [[],[]]
        # arrays = [
                # ["bar", "bar", "baz", "baz", "foo", "foo", "qux", "qux"],
                # ["one", "two", "one", "two", "one", "two", "one", "two"],
                # ["one", "two", "one", "two", "one", "two", "one", "two"]

        # ]
   
        for level in levels:
            sublevels = levels[level]
            df_by_level = pd.DataFrame()
    
            for measure in measures:
                df_splits = pd.DataFrame()
                sp_arr = ["", "", ""]
                for split in splits:
                    df = get_measure_df(measure,level,split)

    
                    try:
                        df = df[df[level].isin(sublevels)]
                    except:
                        st.error("Calculation get_measure_df failed for measure: " + measure)
                        st.write(df.astype(str))
                        st.stop()                   
                    
                    df.insert(0, 'level', level)
                    
                    split_append = ""
                    for s in split:
                        split_append = split_append + s 
                    df = df.rename(columns={level: "sublevel","Total" : "Total_" + split_append})
                    
                    for x in range(0,(df.shape[1]-2)):
                        sp_arr.append(split)
                    
                    if df_splits.empty:
                        df.insert(2, 'measurment', measure)
                        df_splits = df
                    else:
                        df_splits = pd.merge(df_splits, df, how='left', on=["level","sublevel"])

                df_by_level = pd.concat([df_by_level,df_splits])
                if len(arrays[0])==0:
                    arrays[0] = sp_arr
    
            # df_by_level = df_by_level.sort_values(by=['sublevel'])
            df_by_level.reset_index(drop=True, inplace=True)
            
            # df_by_level.loc[df_by_level.shape[0]] = empty_row
    
            # OVDE TREBA DODATI DA UBACI SAMPLE SIZE
            # sampleSizes = data_survey[split].value_counts()
            # sampleSizes['Total'] = sampleSizes.sum()
            # N = pd.DataFrame(data = [sampleSizes], index = ['Sample size'], columns=dfAll.columns)
            # dfAll = pd.concat([N, dfAll])
            
            
            table = pd.concat([table,df_by_level])
            
    
    

        table1 = table.sort_values(by=['level','sublevel','measurment'])
        table1.reset_index(drop=True, inplace=True)
        # st.write(table1)
        table2 = table.sort_values(by=['measurment','level','sublevel'])
        table2.reset_index(drop=True, inplace=True)
        
        st.write(split)
        st.write(table2.astype(str))
        
        tables = {
            "by_level" : table1,
            "by_measure" : table2
        }
    
        if level_number == '1':
            for t in tables:
                table = tables[t]
                empty_dic = {}
                for col in list(table.columns):
                    empty_dic[col] = ["   "]
                empty_df = pd.DataFrame.from_dict(empty_dic)
        
                places_to_insert_empty_df = []
                already_inserted_empty = 0
                for i in range(0,len(table)-1):
                    if i > 0 and table["level"][i] != table["level"][i-1]:
                        places_to_insert_empty_df.append(i+already_inserted_empty)
                        already_inserted_empty+=1
        
                for i in places_to_insert_empty_df:
                    table = pd.concat([table.iloc[:i], empty_df, table.iloc[i:]])
                    table.reset_index(drop=True, inplace=True)
        
        
                arrays[1] = list(table.columns)
                tuples = list(zip(*arrays))
                multi_column_names = pd.MultiIndex.from_tuples(tuples, names=["Var Name", "Var Label"])
                
                table.columns = multi_column_names
                table.reset_index(drop=True, inplace=True)
                tables[t] = table

        dfAll_tables[level_number] = tables
        
        # st.write(tables["by_level"].astype(str))
        # st.write(tables["by_measure"].astype(str))
    # st.write(dfAll_tables)
    return dfAll_tables

def inputEntered ():
    with header2:
        st.markdown("<p style='background-color:#033b6e; color:white'>Data generated for project: <b>" + st.session_state.text_key + "</b>. </br> If you want to change project, just re-enter the number in the input below and press Enter.</p>", unsafe_allow_html=True)
        st.markdown("-------------------------------------------------------------------------------------------")
		
header1 = st.container()
header2 = st.container()
dataset = st.container()

with header1:
    titles = st.title('Hello [user]!')

with header2:

    proj_number = st.text_input("Enter the project number:", value="", autocomplete="on", placeholder= "7-digit project number (ex. 2022126)", on_change=inputEntered, key='text_key')


with dataset:
    # st.write(st.session_state)
    #file uploaders
    #survey_db = st.file_uploader('Upload Survey Database:', type=None, accept_multiple_files=False, key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False)
    #vs_db_files = st.file_uploader('Upload VS Database:', type=None, accept_multiple_files=True, key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False)
    #datamap_json_file = st.file_uploader('Upload JSON Datamap:', type=None, accept_multiple_files=False, key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False)
    #if survey_db is None or datamap_json_file is None or len(vs_db_files)==0:#ovde dodaj i uslov za VS
    if proj_number:
        dataset.empty()
        #datamap to formated json
        dm_json = get_datamap("datamap_json_file")
        datamap = dm_json[0]
        questions_label_text = dm_json[1]
        #Survey data
        surveyFinalData = get_survey_data("survey_db")
        #VS database
        df_vs = get_vs_data("vs_db_files")
        # st.write(df_vs)
        #splits
        st.sidebar.write(pd.DataFrame(questions_label_text, index=None, columns=["questions"]))
        st.sidebar.download_button(label="datamap json", data=str(datamap), file_name="datamap.json",mime='text/csv')
        # st.sidebar.write(datamap)
        # st.stop()
        parameters = {}
        col_measurments,col_splits = st.columns(2)
        
        with col_measurments:
            st.info("Choose measurments:")
            measurments = ["Consideration on total sample","Penetration on total sample","Total Units","Total Value","Share of Total Units","Share of Total Value","Unit Buy Rate (Units per Buyer)","Value Buy Rate(Value per Buyer)"]
            
            parameters["measurments"] = {}
            
            measurments_select_all = st.checkbox("ALL MEASUREMENT")
            for m in measurments:
                if measurments_select_all:
                    parameters["measurments"][m] = st.checkbox(m, value=True)
                else:
                    parameters["measurments"][m] = st.checkbox(m)

        splits_long = {}
        with col_splits:
            st.info("Add splits lvl1:")
            splits_long["1"] =  st.multiselect("Type to search or just scroll:",questions_label_text, key="splits_lvl1")

            st.info("Add splits lvl2:")
            splits_long["2"] =  st.multiselect("Type to search or just scroll:",questions_label_text, key="splits_lvl2")

            st.info("Add splits lvl3:")
            splits_long["3"] =  st.multiselect("Type to search or just scroll:",questions_label_text, key="splits_lvl3")
            if len(splits_long["3"])>0 and len(splits_long["2"])==0:
                st.error("You can't have split level 3, before you define split level 2!!!")
                st.stop()

        splits_final = format_splits(splits_long)
        
        
        col_lev1,col_lev2 = st.columns(2)
    
        with col_lev1:
            st.info("Choose levels:")
            levels = ["SKU","BRAND","SUBBRAND","PRODUCT CATEGORY","PURPOSE","CLIENT","UNIT OF MEASUREMENT","SHELF","KPI1","KPI2","KPI3","KPI4","KPI5","PRODUCT DESCRIPTION 1","PRODUCT DESCRIPTION 2","PRODUCT DESCRIPTION 3","Custom attribute levels"]
            
            parameters["levels"] = {}
            
            temporarly_disabled_levels = ["KPI1","KPI2"]
            
            levels_select_all = st.checkbox("ALL LEVELS")
            for level in levels:
                if level in df_vs.columns:
                    if ( len(list(df_vs[level].unique()))==1 and list(df_vs[level].unique())[0]=="NOT DEFINED" ) or level in temporarly_disabled_levels:
                        parameters["levels"][level] = st.checkbox(level, disabled=True, key="lvl_"+level)
                    elif levels_select_all:
                        parameters["levels"][level] = st.checkbox(level, value=True, key="lvl_"+level)
                    else:
                        parameters["levels"][level] = st.checkbox(level, key="lvl_"+level)
    
        with col_lev2:
            parameters["sublevels"] = {}
    
            if not levels_select_all:
                for level in parameters["levels"]:
                    if parameters["levels"][level]==True:
                        st.info(level)
                        sublevels_select_all = st.checkbox("Select all", key="sh_" + level, value = True)
                        
                        sublevels_list = list(df_vs[level].unique())
                        sublevels_list.sort()
                        if "NOT DEFINED" in sublevels_list:
                            sublevels_list.remove("NOT DEFINED")
                            sublevels_list.append("NOT DEFINED")
    
                        if not sublevels_select_all:
                            parameters["sublevels"][level] = st.multiselect("Type to search or just scroll:",sublevels_list, key="sublevel_"+str(level))
                        elif sublevels_select_all:
                            parameters["sublevels"][level] = sublevels_list
                        
            elif levels_select_all:
                for level in parameters["levels"]:
                    sublevels_list = list(df_vs[level].unique())
                    sublevels_list.sort()
                    parameters["sublevels"][level] = sublevels_list
    
        data_survey = get_df_with_answer_labels(surveyFinalData,uuid_and_split)
    
        shoppingMergedData = pd.merge(data_survey, df_vs, how='left', left_on='uuid', right_on='USER ID')
    
        #STARO
        if st.button("Run calculations - V1"):
            chosen_measures = []
            for m in parameters["measurments"]:
                if parameters["measurments"][m]:
                    chosen_measures.append(m)
    
    
            tables_arr = splitEngine(chosen_measures, splits_final["1"], parameters["sublevels"])
    
            tables = tables_arr[0]
            tables_by_measure = tables_arr[1]
    
    
            with pd.ExcelWriter("final.xlsx") as writer:
                filled_sheet_length_1 = 0
                startrow = 0
                for table in tables:
                    table.to_excel(writer, sheet_name="by_level", startrow=startrow, startcol=0, index=True)
                    startrow = startrow + table.shape[0] + 5
                    filled_sheet_length_1+=startrow
    
                format_tables(writer.book, writer.sheets["by_level"], filled_sheet_length_1)
    
                filled_sheet_length_2 = 0
                startrow_measure = 0
                for level in tables_by_measure:
                    for table in tables_by_measure[level]:
                        tables_by_measure[level][table].to_excel(writer, sheet_name="by_measure", startrow=startrow_measure, startcol=0, index=True)
                        startrow_measure = startrow_measure + tables_by_measure[level][table].shape[0] + 5
                        filled_sheet_length_2+=startrow_measure
    
                format_tables(writer.book, writer.sheets["by_measure"], filled_sheet_length_2)
    
        
    
            wb = load_workbook("final.xlsx")
    
            ws = wb['by_level']
            row_reduced_height = []
            for row in ws.iter_rows():
                if not any(cell.value for cell in row):
                    # ws.delete_rows(row[0].row)
                    if row[0].row - 1 not in row_reduced_height:
                        ws.row_dimensions[row[0].row].height = 0.5
                        row_reduced_height.append(row[0].row)
    
            ws = wb['by_measure']
            row_reduced_height = []
            for row in ws.iter_rows():
                if not any(cell.value for cell in row):
                    # ws.delete_rows(row[0].row)
                    if row[0].row - 1 not in row_reduced_height:
                        ws.row_dimensions[row[0].row].height = 0.5
                        row_reduced_height.append(row[0].row)
    
            wb.save("final.xlsx")
    
                
            with open('final.xlsx', mode = "rb") as f:
                st.download_button('Generate Excel Export', f, file_name= 'Export_' + st.session_state.text_key + '_version_1.xlsx')
    
    
    
        #NOVO
        if st.button("Run calculations - V2"):
            with st.spinner('Wait for it...'):
                time.sleep(5)
            st.success('Done!')
            chosen_measures = []
            for m in parameters["measurments"]:
                if parameters["measurments"][m]:
                    chosen_measures.append(m)
    


            tables = splitEngine2(chosen_measures, splits_final, parameters["sublevels"])

            # for split_level in tables:
                # for t in tables[split_level]:
                    # if t == "by_measure":
                        # st.info(split)
                        # st.write(tables[split_level][t].astype(str))
                    # tables[split_level][t].to_excel(writer, sheet_name=t + split_level)
                    # format_tables(writer.book, writer.sheets[t + split_level], len(tables[split_level][t].index) + 3)
                    
            with pd.ExcelWriter("final_by_measure.xlsx") as writer:

                for split_level in tables:
                    for t in tables[split_level]:
                        if t == "by_measure":
                            tables[split_level][t].to_excel(writer, sheet_name=t + split_level)
                            format_tables(writer.book, writer.sheets[t + split_level], len(tables[split_level][t].index) + 3)
            
            with pd.ExcelWriter("final_by_level.xlsx") as writer:
            
                for split_level in tables:
                    for t in tables[split_level]:
                        if t == "by_measure":
                            tables[split_level][t].to_excel(writer, sheet_name=t + split_level)
                            format_tables(writer.book, writer.sheets[t + split_level], len(tables[split_level][t].index) + 3)
    
            zipObj = ZipFile("sample.zip", "w")
            zipObj.write("final_by_measure.xlsx")
            zipObj.write("final_by_level.xlsx")
            zipObj.close()
            ZipfileDotZip = "sample.zip"
            
            
            with open(ZipfileDotZip, "rb") as f:
                bytes = f.read()
                b64 = base64.b64encode(bytes).decode()
                href = f"<a class='download' href=\"data:file/zip;base64,{b64}\" download='{ZipfileDotZip}.zip'>\
                    <b>Download data for project " + st.session_state.text_key + "</b>\
                </a>"
                st.sidebar.markdown(href, unsafe_allow_html=True)

            
            # wb = load_workbook("final.xlsx")
            # ws = wb['by_level']
            # ws.freeze_panes = ws['A4']
            # ws.auto_filter.ref = "A3:AA3"
            # wb.save("final.xlsx")
    
            # wb = load_workbook("final.xlsx")
            # ws = wb['by_measure']
            # ws.freeze_panes = ws['A4']
            # ws.auto_filter.ref = "A3:AA3"
            # wb.save("final.xlsx")
    
    
            with open('final_by_measure.xlsx', mode = "rb") as f:
                st.download_button('Generate Excel Export', f, file_name= 'Export_' + st.session_state.text_key + '_version_by_measure.xlsx')
                
            with open('final_by_level.xlsx', mode = "rb") as f:
                st.download_button('Generate Excel Export', f, file_name= 'Export_' + st.session_state.text_key + '_version_by_level.xlsx')
